import os
import pandas as pd
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import MinMaxScaler
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.ticker import MaxNLocator

def procesar_archivo(ruta_archivo):
    ext = os.path.splitext(ruta_archivo)[-1].lower()
    
    if ext == ".csv":
        df = pd.read_csv(ruta_archivo)
    elif ext in [".xlsx", ".xls"]:
        df = pd.read_excel(ruta_archivo, engine='openpyxl')

    else:
        raise ValueError("Formato de archivo no soportado")

    features = df[[
        "entrada_minutos",
        "salida_minutos",
        "ausente",
        "ausencias_consecutivas",
        "llegadas_tarde_consecutivas",
        "salidas_tempranas_consecutivas"
    ]]

    scaler = MinMaxScaler()
    X_scaled = scaler.fit_transform(features)

    model = IsolationForest(contamination=0.1, random_state=42)
    df["anomaly_flag"] = model.fit_predict(X_scaled)
    df["anomalia"] = df["anomaly_flag"].map({1: "Normal", -1: "Anomalía"})

    def determinar_motivo(row):
        if row["anomalia"] == "Normal":
            return ""
        motivos = []
        if row["ausencias_consecutivas"] >= 2:
            motivos.append(f"Ausencias consecutivas ({row['ausencias_consecutivas']})")
        if row["llegadas_tarde_consecutivas"] >= 2:
            motivos.append(f"Llegadas tarde consecutivas ({row['llegadas_tarde_consecutivas']})")
        if row["salidas_tempranas_consecutivas"] >= 2:
            motivos.append(f"Salidas tempranas consecutivas ({row['salidas_tempranas_consecutivas']})")
        if not motivos:
            return "Patrón atípico general"
        return ", ".join(motivos)

    df["motivo_anomalia"] = df.apply(determinar_motivo, axis=1)
    df["anomaly_legenda"] = df["anomalia"]
    df.loc[df["anomalia"] == "Anomalía", "anomaly_legenda"] = df["motivo_anomalia"]

    df_export = df.drop(columns=[
        "anomaly_flag",
        "ausencias_consecutivas",
        "llegadas_tarde_consecutivas",
        "salidas_tempranas_consecutivas",
        "entrada_minutos",
        "salida_minutos"
    ])
    df_export.to_excel("resultados.xlsx", index=False)

    wb = load_workbook("resultados.xlsx")
    ws = wb.active
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.freeze_panes = "A2"

    col_anomalia = None
    col_motivo = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == "anomalia":
            col_anomalia = idx + 1
        if cell.value == "motivo_anomalia":
            col_motivo = idx + 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if col_anomalia and row[col_anomalia - 1].value == "Anomalía":
            row[col_anomalia - 1].fill = red_fill
        if col_motivo and row[col_motivo - 1].value:
            row[col_motivo - 1].fill = red_fill

    wb.save("resultados.xlsx")

    df_ref = df.copy()
    X_scaled = scaler.fit_transform(df_ref[[
        "entrada_minutos",
        "salida_minutos",
        "ausente",
        "ausencias_consecutivas",
        "llegadas_tarde_consecutivas",
        "salidas_tempranas_consecutivas"
    ]])
    df_ref["anomalia"] = model.predict(X_scaled)
    df_ref["anomalia"] = df_ref["anomalia"].map({1: "Normal", -1: "Anomalía"})
    df_ref["motivo_anomalia"] = df_ref.apply(determinar_motivo, axis=1)

    def clasificar_anomalia(motivo, anomalia):
        if anomalia != "Anomalía":
            return "Normal"
        if "Llegadas tarde consecutivas" in motivo:
            return "Llegadas tarde consecutivas"
        elif "Salidas tempranas consecutivas" in motivo:
            return "Salidas tempranas consecutivas"
        else:
            return None

    df_ref["anomaly_legenda_limpia"] = df_ref.apply(
        lambda row: clasificar_anomalia(row["motivo_anomalia"], row["anomalia"]), axis=1
    )

    df_grafico = df_ref[df_ref["anomaly_legenda_limpia"].isin([
        "Normal", "Llegadas tarde consecutivas", "Salidas tempranas consecutivas"
    ])]

    df_grafico = df_grafico[
        (df_grafico["entrada_minutos"] >= 420) & (df_grafico["entrada_minutos"] <= 1080) &
        (df_grafico["salida_minutos"] >= 420) & (df_grafico["salida_minutos"] <= 1080)
    ]

    def minutos_a_hora(minutos):
        horas = minutos // 60
        minutos_restantes = minutos % 60
        return f"{int(horas):02}:{int(minutos_restantes):02}"

    ticks = list(range(420, 1101, 60))
    tick_labels = [minutos_a_hora(t) for t in ticks]

    colores_grafico = {
        "Normal": "#4CAF50",
        "Llegadas tarde consecutivas": "#E53935",
        "Salidas tempranas consecutivas": "#1E88E5",
    }

    plt.figure(figsize=(10, 6))
    sns.scatterplot(
        data=df_grafico,
        x="salida_minutos",
        y="entrada_minutos",
        hue="anomaly_legenda_limpia",
        palette=colores_grafico,
        s=80,
        edgecolor="black",
        alpha=0.8
    )
    plt.axhline(480, color="gray", linestyle="--", linewidth=1)
    plt.axvline(1020, color="gray", linestyle="--", linewidth=1)
    plt.title("Dispersión: Entrada vs Salida (Normales y Anomalías de Horario)")
    plt.xlabel("Hora de salida")
    plt.ylabel("Hora de entrada")
    plt.xticks(ticks, tick_labels, rotation=45)
    plt.yticks(ticks, tick_labels)
    plt.grid(True, linestyle="--", alpha=0.5)
    plt.tight_layout()
    plt.savefig("grafico_con_normales_y_anomalias_horario.png")

    anomalias_ausencia = df_ref[df_ref["motivo_anomalia"].str.contains("Ausencias", na=False)]
    conteo_anomalias = anomalias_ausencia["nombre_y_apellido_empleado"].value_counts()

    plt.figure(figsize=(10, 6))
    sns.barplot(x=conteo_anomalias.index, y=conteo_anomalias.values, color="salmon")
    plt.gca().yaxis.set_major_locator(MaxNLocator(integer=True))
    plt.title("Anomalías por Ausencias Consecutivas - por Empleado")
    plt.xlabel("Empleado")
    plt.ylabel("Cantidad de anomalías (ausencia)")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig("grafico_anomalias_ausencia_por_empleado.png")

    empleados_con_anomalias = df[df["anomalia"] == "Anomalía"]["nombre_y_apellido_empleado"].unique().tolist()
    total_anomalias = len(df[df["anomalia"] == "Anomalía"])
    total_registros = len(df)
    
    tabla_json = df_export.fillna("").to_dict(orient="records")

    return empleados_con_anomalias, total_anomalias, total_registros, tabla_json

